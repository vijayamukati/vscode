from bs4 import BeautifulSoup
import openpyxl 
from openpyxl import workbook
import pandas as pd

# Read the HTML file
with open(r'C:\Users\dell\Documents\project\amazon scraping\Amazon.html',encoding="utf8") as f:
   html = f.read()

# Parse the HTML file using BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Find all the divs with the specified class name
divs = soup.find_all('div', {'class': 's-card-container s-overflow-hidden aok-relative puis-wide-grid-style puis-wide-grid-style-t2 puis-include-content-margin puis s-latency-cf-section s-card-border'})

# Create an empty list to store the extracted data
data = []

# Loop through each div and extract the product name, reviews, and price
for div in divs:
    product_name = div.find('span', {'class': 'a-size-medium a-color-base a-text-normal'})
    if product_name:
        product_name = product_name.text.strip()
    else:
        product_name = ''
        
    product_reviews = div.find('span', {'class': 'a-size-base'})
    if product_reviews:
        product_reviews = product_reviews.text.strip()
    else:
        product_reviews = ''
        
    product_price = div.find('span', {'class': 'a-price-whole'})
    if product_price:
        product_price = product_price.text.strip()
    else:
        product_price = ''
    
    # Append the extracted data to the list
    data.append([product_name, product_reviews, product_price])

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the header row
worksheet.cell(row=1, column=1, value='Product Name')
worksheet.cell(row=1, column=2, value='Product Reviews')
worksheet.cell(row=1, column=3, value='Product Price')

# Write the data to the worksheet
for i, row in enumerate(data):
    worksheet.cell(row=i+2, column=1, value=row[0])
    worksheet.cell(row=i+2, column=2, value=row[1])
    worksheet.cell(row=i+2, column=3, value=row[2])

# Save the workbook
workbook.save('Amazon.xlsx')
